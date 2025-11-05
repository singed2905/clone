import pandas as pd
import os
import gc
import psutil
from typing import Dict, List, Tuple, Any, Optional, Iterator, Callable
from datetime import datetime
import threading
import time

class LargeFileProcessor:
    """
    OPTIMIZED HIGH-SPEED processor for large Excel files - Ph∆∞∆°ng √°n A
    Features: Single-workbook streaming, 400+ rows/sec, TL-matching performance
    Smart keylog column (strict): only 'keylog'. If absent, create 'keylog'.
    Font styling for keylog column: Flexio Fx799VN, 11pt, bold, black.
    """
    
    def __init__(self, config: Dict = None):
        self.config = config or {}
        self.processing_cancelled = False
        self.max_memory_mb = 1500
        self.emergency_cleanup = False
        self.max_rows_allowed = 250_000
        self.fast_mode = True
        
    def get_memory_usage(self) -> float:
        try:
            process = psutil.Process()
            return process.memory_info().rss / 1024 / 1024
        except:
            return 0.0
    
    def check_memory_limit(self) -> bool:
        return self.get_memory_usage() > self.max_memory_mb
    
    def emergency_memory_cleanup(self):
        self.emergency_cleanup = True
        gc.collect()
        
    def _enforce_row_limit(self, total_rows: int):
        if total_rows > self.max_rows_allowed:
            raise Exception(
                f"File c√≥ {total_rows:,} d√≤ng, v∆∞·ª£t qu√° gi·ªõi h·∫°n t·ªëi ƒëa {self.max_rows_allowed:,} d√≤ng.\n"
                f"Vui l√≤ng chia nh·ªè file ho·∫∑c l·ªçc b·ªõt d·ªØ li·ªáu tr∆∞·ªõc khi import."
            )
    
    def estimate_optimal_chunksize(self, file_path: str) -> int:
        try:
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            if file_size_mb > 100:
                return 3000
            elif file_size_mb > 50:
                return 5000
            elif file_size_mb > 20:
                return 7000
            else:
                return 10000
        except Exception:
            return 5000
    
    def _get_actual_total_rows(self, file_path: str) -> int:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            actual_total = max(0, (ws.max_row - 1)) if hasattr(ws, 'max_row') and ws.max_row else 0
            wb.close()
            print(f"üìä Actual file dimensions: {actual_total:,} data rows")
            return actual_total
        except Exception as e:
            print(f"‚ö†Ô∏è Could not get actual rows: {e}")
            return 0
    
    def _detect_keylog_column_strict(self, file_path: str) -> Tuple[bool, str, int]:
        """Strict detection: only 'keylog' (case-insensitive). Returns (has_keylog, 'keylog', index or -1)."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()
            for i, cell in enumerate(header_row):
                if cell is not None and str(cell).strip().lower() == 'keylog':
                    print(f"üîç Found strict keylog column at position {i+1}")
                    return True, 'keylog', i
            print("üìù No strict 'keylog' column found. Will create new: 'keylog'")
            return False, 'keylog', -1
        except Exception as e:
            print(f"‚ö†Ô∏è Strict keylog detection failed: {e}")
            return False, 'keylog', -1
    
    def read_excel_streaming_single_workbook(self, file_path: str, chunksize: int = None) -> Iterator[pd.DataFrame]:
        if chunksize is None:
            chunksize = self.estimate_optimal_chunksize(file_path)
        try:
            import openpyxl
            print(f"üöÄ PH∆Ø∆†NG √ÅN A: Single-workbook streaming")
            print(f"üìÅ File: {os.path.basename(file_path)}")
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            max_row = ws.max_row if hasattr(ws, 'max_row') and ws.max_row else 0
            max_col = ws.max_column if hasattr(ws, 'max_column') else 0
            total_rows = max(0, max_row - 1)
            self._enforce_row_limit(total_rows)
            print(f"üìä Dimensions: {total_rows:,} rows √ó {max_col} columns")
            print(f"‚ö° Chunk size: {chunksize:,} rows")
            print(f"üì¶ Estimated chunks: {(total_rows + chunksize - 1) // chunksize}")
            try:
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                columns = [str(cell) if cell is not None else f"Col_{i}" for i, cell in enumerate(header_row)]
            except Exception:
                columns = [f"Col_{i}" for i in range(max_col)]
            current_row = 2
            chunk_count = 0
            while current_row <= max_row and not self.processing_cancelled:
                try:
                    end_row = min(current_row + chunksize - 1, max_row)
                    print(f"‚ö° Reading chunk {chunk_count + 1}: rows {current_row:,}-{end_row:,}")
                    chunk_data = []
                    for row in ws.iter_rows(min_row=current_row, max_row=end_row, values_only=True):
                        row_data = [str(cell) if cell is not None else "" for cell in row]
                        chunk_data.append(row_data)
                    if chunk_data:
                        chunk_df = pd.DataFrame(chunk_data, columns=columns).fillna('')
                        yield chunk_df
                        del chunk_df, chunk_data
                        if chunk_count % 10 == 0 and chunk_count > 0:
                            gc.collect()
                            print(f"üßπ Cleanup checkpoint: Memory {self.get_memory_usage():.1f}MB")
                    current_row = end_row + 1
                    chunk_count += 1
                except Exception as e:
                    print(f"‚ùå Error reading chunk {chunk_count + 1}: {e}")
                    current_row += chunksize
                    continue
            wb.close()
            print(f"‚úÖ Single-workbook streaming completed: {chunk_count} chunks")
        except Exception as e:
            raise Exception(f"L·ªói streaming single-workbook: {str(e)}")
    
    def process_large_excel_fast(self, file_path: str, shape_a: str, shape_b: str,
                                operation: str, dimension_a: str, dimension_b: str,
                                output_path: str, progress_callback: Callable = None) -> Tuple[int, int, str]:
        self.processing_cancelled = False
        success_count = 0
        error_count = 0
        processed_count = 0
        start_time = time.time()
        temp_results_file = f"{output_path}.temp_results"
        from services.geometry.geometry_service import GeometryService
        try:
            print(f"üöÄ PH∆Ø∆†NG √ÅN A - HIGH-SPEED processing: {os.path.basename(file_path)}")
            has_keylog, keylog_col_name, keylog_col_index = self._detect_keylog_column_strict(file_path)
            total_rows = self._get_actual_total_rows(file_path)
            self._enforce_row_limit(total_rows)
            service = GeometryService(self.config)
            service.set_current_shapes(shape_a, shape_b)
            service.set_kich_thuoc(dimension_a, dimension_b)
            service.set_current_operation(operation)
            chunk_size = self.estimate_optimal_chunksize(file_path)
            print(f"‚ö° Optimized chunk size: {chunk_size:,} rows")
            print(f"üéØ Target: {total_rows:,} rows at 400+ rows/sec")
            results_buffer = []
            buffer_size = 5000
            chunk_count = 0
            last_speed_check = time.time()
            for chunk_df in self.read_excel_streaming_single_workbook(file_path, chunk_size):
                if self.processing_cancelled:
                    break
                chunk_count += 1
                chunk_start = time.time()
                chunk_results = []
                for index, row in chunk_df.iterrows():
                    try:
                        if self.processing_cancelled:
                            break
                        data_a = self._extract_shape_data_fast(row, shape_a, 'A')
                        data_b = self._extract_shape_data_fast(row, shape_b, 'B') if shape_b else {}
                        if processed_count % 10000 == 0 and processed_count > 0:
                            service = GeometryService(self.config)
                            service.set_current_shapes(shape_a, shape_b)
                            service.set_kich_thuoc(dimension_a, dimension_b)
                            service.set_current_operation(operation)
                        service.thuc_thi_tat_ca(data_a, data_b)
                        result = service.generate_final_result()
                        chunk_results.append(result)
                        success_count += 1
                    except Exception as e:
                        chunk_results.append(f"L·ªñI: {str(e)}")
                        error_count += 1
                    processed_count += 1
                results_buffer.extend(chunk_results)
                if len(results_buffer) >= buffer_size:
                    self._write_results_buffer_fast(temp_results_file, results_buffer)
                    results_buffer = []
                chunk_time = time.time() - chunk_start
                chunk_speed = len(chunk_df) / chunk_time if chunk_time >= 0.5 else None
                current_time = time.time()
                elapsed = current_time - start_time
                avg_speed = processed_count / elapsed if elapsed > 0 else 0
                if progress_callback and processed_count % 100 == 0:
                    processed_display = min(processed_count, total_rows)
                    progress_percent = (processed_display / total_rows) * 100 if total_rows > 0 else 0
                    progress_callback(progress_percent, processed_display, total_rows, error_count)
                if current_time - last_speed_check >= 5:
                    processed_display = min(processed_count, total_rows)
                    progress_percent = (processed_display / total_rows) * 100 if total_rows > 0 else 0
                    remaining_rows = total_rows - processed_display
                    eta_seconds = remaining_rows / max(avg_speed, 1e-6) if remaining_rows > 0 else 0
                    eta_minutes = int(eta_seconds // 60)
                    eta_secs = int(eta_seconds % 60)
                    eta_str = f"{eta_minutes:02d}:{eta_secs:02d}"
                    if chunk_speed is not None:
                        speed_display = f"üî• Speed: {avg_speed:.0f} rows/sec (avg) | {chunk_speed:.0f} rows/sec (current)"
                    else:
                        speed_display = f"üî• Speed: {avg_speed:.0f} rows/sec (avg)"
                    print(f"{speed_display} | Progress: {processed_display:,}/{total_rows:,} ({progress_percent:.1f}%) | ETA: {eta_str}")
                    last_speed_check = current_time
                if chunk_count % 5 == 0 and self.check_memory_limit():
                    print(f"‚ö†Ô∏è Memory: {self.get_memory_usage():.1f}MB - Quick cleanup")
                    gc.collect()
                del chunk_df
            if results_buffer:
                self._write_results_buffer_fast(temp_results_file, results_buffer)
            print("üîß Creating final Excel file with strict keylog + Flexio font...")
            final_output = self._create_excel_with_smart_keylog(file_path, temp_results_file, output_path, 
                                                               has_keylog, keylog_col_name, keylog_col_index)
            total_time = time.time() - start_time
            final_speed = processed_count / total_time if total_time > 0 else 0
            print(f"üèÅ PH∆Ø∆†NG √ÅN A COMPLETED!")
            print(f"‚ö° Final speed: {final_speed:.0f} rows/sec (Target: 400+ rows/sec)")
            print(f"üìä Total: {processed_count:,} rows in {total_time:.1f}s")
            print(f"‚úÖ Success: {success_count:,} | ‚ùå Errors: {error_count:,}")
            return success_count, error_count, final_output
        except Exception as e:
            raise Exception(f"L·ªói x·ª≠ l√Ω PH∆Ø∆†NG √ÅN A: {str(e)}")
        finally:
            try:
                if os.path.exists(temp_results_file):
                    os.remove(temp_results_file)
                    print(f"üßπ Cleaned up temp file: {os.path.basename(temp_results_file)}")
            except Exception as cleanup_err:
                print(f"‚ö†Ô∏è Could not remove temp file: {cleanup_err}")
    
    def _create_excel_with_smart_keylog(self, original_file: str, temp_results_file: str, 
                                       output_path: str, has_keylog: bool, keylog_col_name: str, 
                                       keylog_col_index: int) -> str:
        try:
            from openpyxl.styles import Font
            print(f"‚ö° SMART KEYLOG Excel creation (strict 'keylog' + Flexio font)...")
            all_results = self._read_temp_results_fast(temp_results_file)
            # Always use 'keylog' as the column name
            keylog_col_name = 'keylog'
            if has_keylog:
                print(f"üìù Using existing 'keylog' column (position {keylog_col_index + 1 if keylog_col_index>=0 else '?'} )")
            else:
                print(f"üìù Creating new 'keylog' column")
            try:
                original_df = pd.read_excel(original_file, dtype=str, keep_default_na=False, engine='openpyxl')
                results_to_add = all_results[:len(original_df)]
                if len(results_to_add) < len(original_df):
                    results_to_add.extend([''] * (len(original_df) - len(results_to_add)))
                if 'keylog' in [c.strip().lower() for c in original_df.columns]:
                    # Update by exact name (case-insensitive match)
                    # Find exact column name to preserve original casing if any
                    exact_name = next((c for c in original_df.columns if c.strip().lower() == 'keylog'), 'keylog')
                    original_df[exact_name] = results_to_add
                    applied_name = exact_name
                else:
                    original_df['keylog'] = results_to_add
                    applied_name = 'keylog'
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    original_df.to_excel(writer, sheet_name='Results', index=False)
                    worksheet = writer.sheets['Results']
                    # Apply Flexio Fx799VN 11 bold black to keylog column
                    header_cells = next(worksheet.iter_rows(min_row=1, max_row=1))
                    keylog_col_idx = None
                    for idx, cell in enumerate(header_cells, start=1):
                        if str(cell.value).strip().lower() == 'keylog':
                            keylog_col_idx = idx
                            break
                    if keylog_col_idx:
                        keylog_font = Font(name="Flexio Fx799VN", size=11, bold=True, color="000000")
                        max_row = worksheet.max_row
                        for row in range(1, max_row + 1):
                            worksheet.cell(row=row, column=keylog_col_idx).font = keylog_font
                    # Minimal auto-width for readability
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = min(len(str(cell.value)), 40)
                            except:
                                pass
                        worksheet.column_dimensions[column_letter].width = max(max_length + 2, 10)
                print(f"‚úÖ SMART KEYLOG Excel creation completed!")
                return output_path
            except Exception as pandas_error:
                print(f"‚ö†Ô∏è Pandas method failed: {pandas_error}")
                return self._create_excel_openpyxl_smart_keylog(original_file, all_results, output_path, 
                                                               has_keylog, 'keylog', keylog_col_index)
        except Exception as e:
            raise Exception(f"L·ªói t·∫°o Excel SMART KEYLOG: {str(e)}")
    
    def _create_excel_openpyxl_smart_keylog(self, original_file: str, results: List[str], output_path: str,
                                           has_keylog: bool, keylog_col_name: str, keylog_col_index: int) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font
            print("üîÑ Using openpyxl smart keylog method (strict 'keylog')...")
            source_wb = openpyxl.load_workbook(original_file, read_only=True, data_only=True)
            source_ws = source_wb.active
            output_wb = openpyxl.Workbook()
            output_ws = output_wb.active
            output_ws.title = "Results"
            header_row = next(source_ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header_list = list(header_row)
            # Determine if 'keylog' exists in header (strict)
            target_col_index = None
            for i, cell in enumerate(header_list):
                if cell is not None and str(cell).strip().lower() == 'keylog':
                    target_col_index = i
                    break
            if target_col_index is None:
                # Create new keylog column at the end
                final_header = header_list + ['keylog']
                target_col_index = len(header_list)
                print(f"üìù Will create new column 'keylog' at position {target_col_index + 1}")
            else:
                final_header = header_list
                print(f"üìù Will update existing column 'keylog' at position {target_col_index + 1}")
            # Write header
            for col_idx, header_cell in enumerate(final_header):
                cell_value = str(header_cell) if header_cell is not None else ""
                output_ws.cell(row=1, column=col_idx + 1, value=cell_value)
            # Write data rows with 'keylog'
            row_count = 2
            result_idx = 0
            for data_row in source_ws.iter_rows(min_row=2, values_only=True):
                if result_idx >= len(results):
                    break
                data_list = list(data_row)
                if len(data_list) < len(final_header):
                    data_list += [""] * (len(final_header) - len(data_list))
                data_list[target_col_index] = results[result_idx]
                for col_idx, cell_value in enumerate(data_list):
                    value = str(cell_value) if cell_value is not None else ""
                    output_ws.cell(row=row_count, column=col_idx + 1, value=value)
                row_count += 1
                result_idx += 1
                if row_count % 5000 == 0:
                    gc.collect()
            # Apply Flexio font to keylog column
            keylog_font = Font(name="Flexio Fx799VN", size=11, bold=True, color="000000")
            col = target_col_index + 1
            max_row = output_ws.max_row
            for row in range(1, max_row + 1):
                output_ws.cell(row=row, column=col).font = keylog_font
            source_wb.close()
            output_wb.save(output_path)
            output_wb.close()
            print(f"‚úÖ Openpyxl smart keylog method completed!")
            return output_path
        except Exception as e:
            raise Exception(f"L·ªói openpyxl smart keylog fallback: {str(e)}")
    
    def _extract_shape_data_fast(self, row: pd.Series, shape_type: str, group: str) -> Dict:
        data_dict = {}
        if not shape_type:
            return data_dict
        if group == 'A':
            if shape_type == "ƒêi·ªÉm":
                data_dict['point_input'] = str(row.get('data_A', '')).strip()
            elif shape_type == "ƒê∆∞·ªùng th·∫≥ng":
                data_dict['line_A1'] = str(row.get('d_P_data_A', '')).strip()
                data_dict['line_X1'] = str(row.get('d_V_data_A', '')).strip()
            elif shape_type == "M·∫∑t ph·∫≥ng":
                data_dict['plane_a'] = str(row.get('P1_a', '')).strip()
                data_dict['plane_b'] = str(row.get('P1_b', '')).strip()
                data_dict['plane_c'] = str(row.get('P1_c', '')).strip()
                data_dict['plane_d'] = str(row.get('P1_d', '')).strip()
            elif shape_type == "ƒê∆∞·ªùng tr√≤n":
                data_dict['circle_center'] = str(row.get('C_data_I1', '')).strip()
                data_dict['circle_radius'] = str(row.get('C_data_R1', '')).strip()
            elif shape_type == "M·∫∑t c·∫ßu":
                data_dict['sphere_center'] = str(row.get('S_data_I1', '')).strip()
                data_dict['sphere_radius'] = str(row.get('S_data_R1', '')).strip()
        else:
            if shape_type == "ƒêi·ªÉm":
                data_dict['point_input'] = str(row.get('data_B', '')).strip()
            elif shape_type == "ƒê∆∞·ªùng th·∫≥ng":
                data_dict['line_A2'] = str(row.get('d_P_data_B', '')).strip()
                data_dict['line_X2'] = str(row.get('d_V_data_B', '')).strip()
            elif shape_type == "M·∫∑t ph·∫≥ng":
                data_dict['plane_a'] = str(row.get('P2_a', '')).strip()
                data_dict['plane_b'] = str(row.get('P2_b', '')).strip()
                data_dict['plane_c'] = str(row.get('P2_c', '')).strip()
                data_dict['plane_d'] = str(row.get('P2_d', '')).strip()
            elif shape_type == "ƒê∆∞·ªùng tr√≤n":
                data_dict['circle_center'] = str(row.get('C_data_I2', '')).strip()
                data_dict['circle_radius'] = str(row.get('C_data_R2', '')).strip()
            elif shape_type == "M·∫∑t c·∫ßu":
                data_dict['sphere_center'] = str(row.get('S_data_I2', '')).strip()
                data_dict['sphere_radius'] = str(row.get('S_data_R2', '')).strip()
        return data_dict
    
    def _write_results_buffer_fast(self, temp_file: str, results: List[str]):
        try:
            mode = 'a' if os.path.exists(temp_file) else 'w'
            with open(temp_file, mode, encoding='utf-8', buffering=16384) as f:
                f.write('\n'.join(results) + '\n')
        except Exception as e:
            print(f"‚ö†Ô∏è Warning: Fast buffer write failed: {e}")
    
    def _read_temp_results_fast(self, temp_file: str) -> List[str]:
        try:
            with open(temp_file, 'r', encoding='utf-8', buffering=16384) as f:
                return f.read().strip().split('\n')
        except Exception:
            return []
    
    def _create_excel_direct_fast(self, original_file: str, temp_results_file: str, output_path: str) -> str:
        # Redirect to smart keylog creator with strict behavior
        return self._create_excel_with_smart_keylog(original_file, temp_results_file, output_path,
                                                   *self._detect_keylog_column_strict(original_file))
    
    def _create_excel_openpyxl_fast(self, original_file: str, results: List[str], output_path: str) -> str:
        has_keylog, keylog_col_name, keylog_col_index = self._detect_keylog_column_strict(original_file)
        return self._create_excel_openpyxl_smart_keylog(original_file, results, output_path,
                                                       has_keylog, keylog_col_name, keylog_col_index)
    
    def process_large_excel_safe(self, file_path: str, shape_a: str, shape_b: str,
                                operation: str, dimension_a: str, dimension_b: str,
                                output_path: str, progress_callback: Callable = None) -> Tuple[int, int, str]:
        return self.process_large_excel_fast(file_path, shape_a, shape_b, operation, 
                                           dimension_a, dimension_b, output_path, progress_callback)
    
    def validate_large_file_structure(self, file_path: str, shape_a: str, shape_b: str = None) -> Dict[str, Any]:
        try:
            actual_rows = self._get_actual_total_rows(file_path)
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            header_df = pd.read_excel(file_path, nrows=0, engine='openpyxl')
            columns = header_df.columns.tolist()
            has_keylog, keylog_col_name, keylog_col_index = self._detect_keylog_column_strict(file_path)
            over_limit = actual_rows > self.max_rows_allowed
            required_columns_A = self._get_required_columns(shape_a, 'A')
            required_columns_B = self._get_required_columns(shape_b, 'B') if shape_b else []
            missing_columns = [col for col in (required_columns_A + required_columns_B) if col not in columns]
            return {
                'valid': len(missing_columns) == 0 and not over_limit,
                'file_size_mb': file_size_mb,
                'actual_rows': actual_rows,
                'columns': columns,
                'missing_columns': missing_columns,
                'has_keylog_column': has_keylog,
                'keylog_column_name': keylog_col_name,
                'keylog_column_index': keylog_col_index,
                'recommended_chunk_size': self.estimate_optimal_chunksize(file_path),
                'max_rows_allowed': self.max_rows_allowed,
                'over_row_limit': over_limit,
                'validation_method': 'ph∆∞∆°ng_√°n_A_strict_keylog',
                'warning': f'File v∆∞·ª£t qu√° gi·ªõi h·∫°n {self.max_rows_allowed:,} d√≤ng' if over_limit else ''
            }
        except Exception as e:
            return {'valid': False, 'error': f'L·ªói ki·ªÉm tra file: {str(e)}'}
    
    def _get_required_columns(self, shape: str, group: str) -> List[str]:
        if not shape:
            return []
        if group == 'A':
            mapping = {
                "ƒêi·ªÉm": ["data_A"],
                "ƒê∆∞·ªùng th·∫≥ng": ["d_P_data_A", "d_V_data_A"],
                "M·∫∑t ph·∫≥ng": ["P1_a", "P1_b", "P1_c", "P1_d"],
                "ƒê∆∞·ªùng tr√≤n": ["C_data_I1", "C_data_R1"],
                "M·∫∑t c·∫ßu": ["S_data_I1", "S_data_R1"]
            }
        else:
            mapping = {
                "ƒêi·ªÉm": ["data_B"],
                "ƒê∆∞·ªùng th·∫≥ng": ["d_P_data_B", "d_V_data_B"],
                "M·∫∑t ph·∫≥ng": ["P2_a", "P2_b", "P2_c", "P2_d"],
                "ƒê∆∞·ªùng tr√≤n": ["C_data_I2", "C_data_R2"],
                "M·∫∑t c·∫ßu": ["S_data_I2", "S_data_R2"]
            }
        return mapping.get(shape, [])
    
    def get_processing_statistics(self) -> Dict[str, Any]:
        return {
            'memory_usage_mb': self.get_memory_usage(),
            'memory_limit_mb': self.max_memory_mb,
            'memory_usage_percent': (self.get_memory_usage() / self.max_memory_mb) * 100,
            'emergency_cleanup_triggered': self.emergency_cleanup,
            'processing_cancelled': self.processing_cancelled,
            'recommended_max_chunksize': 5000 if self.get_memory_usage() > 800 else 7000,
            'max_rows_allowed': self.max_rows_allowed,
            'processing_method': 'ph∆∞∆°ng_√°n_A_strict_keylog_flexio',
            'target_speed_rows_per_sec': 400,
            'optimizations': ['single_workbook_open', 'strict_keylog', 'flexio_font', 'large_chunks', 'minimal_gc', 'batch_io']
        }