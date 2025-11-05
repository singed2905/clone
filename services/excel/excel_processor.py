import pandas as pd
import json
import os
import openpyxl
from typing import Dict, List, Tuple, Any, Optional
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import re
from datetime import datetime
from .large_file_processor import LargeFileProcessor

class ExcelProcessor:
    """Excel Processor for ConvertKeylogApp - Enhanced with large file support"""
    
    def __init__(self, config: Dict = None):
        self.config = config or {}
        self.mapping = self._load_mapping()
        self.large_file_processor = LargeFileProcessor(config)  # NEW: Large file handler
        self.large_file_threshold_mb = 20  # Files > 20MB use large file processor
        self.large_file_threshold_rows = 50000  # Files > 50k rows use large file processor
    
    def _load_mapping(self) -> Dict:
        """Load Excel mapping configuration from config or fallback"""
        try:
            # Try to load from new config structure
            if self.config and 'geometry' in self.config:
                geometry_config = self.config['geometry']
                if 'excel_mapping' in geometry_config:
                    return geometry_config['excel_mapping']
            
            # Fallback: try to load from separate file
            mapping_file = "config/geometry_mode/geometry_excel_mapping.json"
            if os.path.exists(mapping_file):
                with open(mapping_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
                    
        except Exception as e:
            print(f"Warning: Could not load Excel mapping: {e}")
        
        # Default mapping fallback
        return self._get_default_mapping()
    
    def _get_default_mapping(self) -> Dict:
        """Default Excel mapping matching TL structure"""
        return {
            "group_a_mapping": {
                "Điểm": {
                    "required_columns": ["data_A"],
                    "columns": {
                        "point_input": {"excel_column": "data_A"}
                    }
                },
                "Đường thẳng": {
                    "required_columns": ["d_P_data_A", "d_V_data_A"],
                    "columns": {
                        "line_A1": {"excel_column": "d_P_data_A"},
                        "line_X1": {"excel_column": "d_V_data_A"}
                    }
                },
                "Mặt phẳng": {
                    "required_columns": ["P1_a", "P1_b", "P1_c", "P1_d"],
                    "columns": {
                        "plane_a": {"excel_column": "P1_a"},
                        "plane_b": {"excel_column": "P1_b"},
                        "plane_c": {"excel_column": "P1_c"},
                        "plane_d": {"excel_column": "P1_d"}
                    }
                },
                "Đường tròn": {
                    "required_columns": ["C_data_I1", "C_data_R1"],
                    "columns": {
                        "circle_center": {"excel_column": "C_data_I1"},
                        "circle_radius": {"excel_column": "C_data_R1"}
                    }
                },
                "Mặt cầu": {
                    "required_columns": ["S_data_I1", "S_data_R1"],
                    "columns": {
                        "sphere_center": {"excel_column": "S_data_I1"},
                        "sphere_radius": {"excel_column": "S_data_R1"}
                    }
                }
            },
            "group_b_mapping": {
                "Điểm": {
                    "required_columns": ["data_B"],
                    "columns": {
                        "point_input": {"excel_column": "data_B"}
                    }
                },
                "Đường thẳng": {
                    "required_columns": ["d_P_data_B", "d_V_data_B"],
                    "columns": {
                        "line_A2": {"excel_column": "d_P_data_B"},
                        "line_X2": {"excel_column": "d_V_data_B"}
                    }
                },
                "Mặt phẳng": {
                    "required_columns": ["P2_a", "P2_b", "P2_c", "P2_d"],
                    "columns": {
                        "plane_a": {"excel_column": "P2_a"},
                        "plane_b": {"excel_column": "P2_b"},
                        "plane_c": {"excel_column": "P2_c"},
                        "plane_d": {"excel_column": "P2_d"}
                    }
                },
                "Đường tròn": {
                    "required_columns": ["C_data_I2", "C_data_R2"],
                    "columns": {
                        "circle_center": {"excel_column": "C_data_I2"},
                        "circle_radius": {"excel_column": "C_data_R2"}
                    }
                },
                "Mặt cầu": {
                    "required_columns": ["S_data_I2", "S_data_R2"],
                    "columns": {
                        "sphere_center": {"excel_column": "S_data_I2"},
                        "sphere_radius": {"excel_column": "S_data_R2"}
                    }
                }
            }
        }
    
    def is_large_file(self, file_path: str) -> Tuple[bool, Dict[str, Any]]:
        """Check if file is too large for normal processing"""
        try:
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            
            # Quick row count estimation using openpyxl
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            estimated_rows = ws.max_row - 1 if hasattr(ws, 'max_row') else 0
            wb.close()
            
            is_large = (file_size_mb > self.large_file_threshold_mb or 
                       estimated_rows > self.large_file_threshold_rows)
            
            return is_large, {
                'file_size_mb': file_size_mb,
                'estimated_rows': estimated_rows,
                'recommended_processor': 'large_file' if is_large else 'normal',
                'recommended_chunk_size': self.large_file_processor.estimate_optimal_chunksize(file_path)
            }
            
        except Exception as e:
            return False, {'error': f'Không thể phân tích file: {str(e)}'}
    
    def read_excel_data(self, file_path: str) -> pd.DataFrame:
        """Read Excel file and normalize data - with large file detection"""
        try:
            # Check if this is a large file
            is_large, file_info = self.is_large_file(file_path)
            
            if is_large:
                raise Exception(
                    f"File quá lớn cho phương thức thông thường!\n"
                    f"Kích thước: {file_info.get('file_size_mb', 0):.1f}MB\n"
                    f"Dòng ước tính: {file_info.get('estimated_rows', 0):,}\n\n"
                    f"Vui lòng sử dụng chế độ xử lý file lớn."
                )
            
            df = pd.read_excel(file_path)
            # Normalize column names (remove extra spaces)
            df.columns = df.columns.str.strip()
            return df
        except Exception as e:
            raise Exception(f"Không thể đọc file Excel: {str(e)}")
    
    def validate_excel_structure(self, df: pd.DataFrame, shape_a: str, shape_b: str = None) -> Tuple[bool, List[str]]:
        """Validate Excel structure against selected shapes"""
        missing_columns = []

        # Check Group A columns
        if shape_a in self.mapping['group_a_mapping']:
            required_cols = self.mapping['group_a_mapping'][shape_a]['required_columns']
            for col in required_cols:
                if col not in df.columns:
                    missing_columns.append(f"Nhóm A - {col}")

        # Check Group B columns  
        if shape_b and shape_b in self.mapping['group_b_mapping']:
            required_cols = self.mapping['group_b_mapping'][shape_b]['required_columns']
            for col in required_cols:
                if col not in df.columns:
                    missing_columns.append(f"Nhóm B - {col}")

        return len(missing_columns) == 0, missing_columns
    
    def validate_large_file_structure(self, file_path: str, shape_a: str, shape_b: str = None) -> Dict[str, Any]:
        """Validate large file structure without loading entire file"""
        return self.large_file_processor.validate_large_file_structure(file_path, shape_a, shape_b)
    
    def extract_shape_data(self, row: pd.Series, shape_type: str, group: str) -> Dict:
        """Extract data for specific shape from Excel row"""
        if group == 'A':
            shape_mapping = self.mapping['group_a_mapping'].get(shape_type, {})
        else:
            shape_mapping = self.mapping['group_b_mapping'].get(shape_type, {})

        data_dict = {}
        for field, config in shape_mapping.get('columns', {}).items():
            excel_column = config.get('excel_column')
            if excel_column and excel_column in row.index:
                value = row[excel_column]
                # Convert to string, handle NaN values
                if pd.isna(value):
                    data_dict[field] = ""
                else:
                    data_dict[field] = str(value).strip()
            else:
                data_dict[field] = ""

        return data_dict
    
    def process_large_excel_file(self, file_path: str, shape_a: str, shape_b: str,
                                operation: str, dimension_a: str, dimension_b: str,
                                output_path: str, progress_callback: callable = None) -> Tuple[int, int, str]:
        """Process large Excel files using specialized processor"""
        try:
            print(f"🔥 Switching to LARGE FILE MODE for: {os.path.basename(file_path)}")
            
            return self.large_file_processor.process_large_excel_safe(
                file_path, shape_a, shape_b, operation, dimension_a, dimension_b,
                output_path, progress_callback
            )
            
        except Exception as e:
            raise Exception(f"Lỗi xử lý file lớn: {str(e)}")
    
    def export_results(self, original_df: pd.DataFrame, encoded_results: List[str], output_path: str) -> str:
        """Export results with Excel formatting"""
        try:
            result_df = original_df.copy()

            # Find or create keylog column
            keylog_column = None
            for col in result_df.columns:
                if col.strip().lower() == 'keylog':
                    keylog_column = col
                    break

            # Add results
            if keylog_column:
                result_df[keylog_column] = encoded_results
            else:
                result_df['Kết quả mã hóa'] = encoded_results

            # Ensure directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # Export with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Results')

                # Format the worksheet
                worksheet = writer.sheets['Results']
                self._format_results_worksheet(worksheet, result_df, keylog_column)

            return output_path

        except Exception as e:
            raise Exception(f"Không thể xuất file kết quả: {str(e)}")
    
    def _format_results_worksheet(self, worksheet, df, keylog_column=None):
        """Format Excel worksheet with colors and fonts"""
        try:
            # Header formatting
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')

            # Data formatting
            data_font = Font(name='Arial', size=10)
            result_font = Font(name='Arial', size=10, bold=True, color='2E7D32')

            # Apply header formatting
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

            # Find keylog column index
            keylog_col_name = keylog_column if keylog_column else 'Kết quả mã hóa'
            keylog_col_idx = None
            for idx, col_name in enumerate(df.columns):
                if keylog_col_name in str(col_name):
                    keylog_col_idx = idx
                    break

            # Apply data formatting - limit to first 10k rows for performance
            max_format_rows = min(len(df) + 2, 10000)
            for row in range(2, max_format_rows):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Special formatting for result column
                    if keylog_col_idx is not None and col == keylog_col_idx + 1:
                        cell.font = result_font
                    else:
                        cell.font = data_font

            # Auto-adjust column widths - limited for performance
            for col_idx, column in enumerate(worksheet.columns):
                if col_idx > 20:  # Limit to first 20 columns
                    break
                    
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                # Check only first 100 rows for performance
                for i, cell in enumerate(column):
                    if i > 100:
                        break
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            print(f"Warning: Could not format worksheet: {e}")
    
    def get_total_rows(self, file_path: str) -> int:
        """Get total number of rows in Excel file - optimized for large files"""
        try:
            # For potentially large files, use openpyxl for efficiency
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            
            if file_size_mb > 10:  # Large file - use openpyxl
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                total_rows = ws.max_row - 1 if hasattr(ws, 'max_row') else 0
                wb.close()
                return total_rows
            else:
                # Small file - use pandas
                df = pd.read_excel(file_path)
                return len(df)
        except Exception:
            return 0
    
    def read_excel_data_chunked(self, file_path: str, chunksize: int = 1000):
        """Read Excel data in chunks - Enhanced for large files"""
        try:
            # Check if we need large file processing
            is_large, file_info = self.is_large_file(file_path)
            
            if is_large:
                # Use streaming processor for very large files
                print(f"🔥 Large file detected - using streaming processor")
                return self.large_file_processor.read_excel_streaming(file_path, chunksize)
            else:
                # Use pandas chunking for smaller files
                df = pd.read_excel(file_path)
                # Yield chunks
                for i in range(0, len(df), chunksize):
                    yield df.iloc[i:i + chunksize]
        except Exception as e:
            raise Exception(f"Không thể đọc file Excel theo chunk: {str(e)}")
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get information about Excel file - Enhanced with large file detection"""
        try:
            # Check if large file first
            is_large, large_file_info = self.is_large_file(file_path)
            
            if is_large:
                # Use openpyxl for large file info
                return self._get_large_file_info(file_path, large_file_info)
            else:
                # Use pandas for normal files
                df = self.read_excel_data(file_path)
                file_name = os.path.basename(file_path)

                return {
                    'file_name': file_name,
                    'total_rows': len(df),
                    'total_columns': len(df.columns),
                    'columns': list(df.columns),
                    'file_size': os.path.getsize(file_path),
                    'file_size_mb': os.path.getsize(file_path) / (1024 * 1024),
                    'is_large_file': False,
                    'first_few_rows': df.head(3).to_dict('records') if len(df) > 0 else []
                }
        except Exception as e:
            raise Exception(f"Không thể đọc thông tin file: {str(e)}")
    
    def _get_large_file_info(self, file_path: str, large_file_info: Dict) -> Dict[str, Any]:
        """Get file info for large files without loading data"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            
            # Get header
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            columns = [str(cell) for cell in header_row if cell is not None]
            
            # Get sample rows (first 3 data rows)
            sample_rows = []
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=5, values_only=True)):
                if i >= 3:
                    break
                row_dict = {col: str(cell) if cell is not None else "" for col, cell in zip(columns, row)}
                sample_rows.append(row_dict)
            
            wb.close()
            
            return {
                'file_name': os.path.basename(file_path),
                'total_rows': large_file_info.get('estimated_rows', 0),
                'total_columns': len(columns),
                'columns': columns,
                'file_size': os.path.getsize(file_path),
                'file_size_mb': large_file_info.get('file_size_mb', 0),
                'is_large_file': True,
                'recommended_chunk_size': large_file_info.get('recommended_chunk_size', 500),
                'first_few_rows': sample_rows,
                'warning': 'File lớn - khuyến nghị dùng xử lý chunked'
            }
            
        except Exception as e:
            raise Exception(f"Không thể phân tích file lớn: {str(e)}")
    
    def create_geometry_template(self, shape_a: str, shape_b: str, output_path: str) -> str:
        """Create Excel template for geometry data input"""
        try:
            # Get required columns
            template_data = {}
            
            # Add columns for shape A
            if shape_a in self.mapping['group_a_mapping']:
                required_cols = self.mapping['group_a_mapping'][shape_a]['required_columns']
                for col in required_cols:
                    template_data[col] = self._get_sample_data(shape_a, col)
            
            # Add columns for shape B
            if shape_b and shape_b in self.mapping['group_b_mapping']:
                required_cols = self.mapping['group_b_mapping'][shape_b]['required_columns']
                for col in required_cols:
                    template_data[col] = self._get_sample_data(shape_b, col)
            
            # Add keylog column
            template_data['keylog'] = [''] * len(next(iter(template_data.values())))
            
            # Create DataFrame
            df = pd.DataFrame(template_data)
            
            # Export template
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Template')
                
                # Format template
                worksheet = writer.sheets['Template']
                self._format_template_worksheet(worksheet, df, shape_a, shape_b)
            
            return output_path
            
        except Exception as e:
            raise Exception(f"Không thể tạo template: {str(e)}")
    
    def _get_sample_data(self, shape: str, column: str) -> List[str]:
        """Generate sample data for template"""
        if shape == "Điểm":
            return ['1,2', '3,4', '5,6', '0,0']
        elif shape == "Đường thẳng":
            if 'P_data' in column:
                return ['0,0,0', '1,1,1', '2,3,4', '0,1,0']
            else:  # V_data
                return ['1,0,0', '0,1,0', '1,1,1', '0,0,1']
        elif shape == "Mặt phẳng":
            if column.endswith('_a'):
                return ['1', '2', '1', '0']
            elif column.endswith('_b'):
                return ['1', '1', '2', '1']
            elif column.endswith('_c'):
                return ['1', '3', '1', '0']
            else:  # _d
                return ['0', '5', '6', '1']
        elif shape == "Đường tròn":
            if 'I' in column:  # Center
                return ['0,0', '1,1', '2,3', '0,5']
            else:  # Radius
                return ['5', '3', '2', '10']
        elif shape == "Mặt cầu":
            if 'I' in column:  # Center
                return ['0,0,0', '1,1,1', '2,3,4', '0,0,5']
            else:  # Radius
                return ['5', '3', '2', '10']
        
        return ['', '', '', '']
    
    def _format_template_worksheet(self, worksheet, df, shape_a, shape_b):
        """Format template worksheet with instructions"""
        try:
            # Header formatting
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
            
            # Apply header formatting
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            print(f"Warning: Could not format template: {e}")
    
    def validate_data_quality(self, df: pd.DataFrame, shape_a: str, shape_b: str = None) -> Dict[str, Any]:
        """Validate data quality in Excel file - optimized for large files"""
        # For large DataFrames, limit validation to first 1000 rows
        sample_size = min(1000, len(df))
        sample_df = df.head(sample_size) if len(df) > 1000 else df
        
        quality_info = {
            'valid': True,
            'total_rows': len(df),
            'sample_size': sample_size,
            'rows_with_data': 0,
            'rows_with_errors': 0,
            'missing_columns': [],
            'data_issues': [],
            'is_sample_validation': len(df) > 1000
        }
        
        # Check structure first
        is_valid, missing_cols = self.validate_excel_structure(df, shape_a, shape_b)
        if not is_valid:
            quality_info['valid'] = False
            quality_info['missing_columns'] = missing_cols
            return quality_info
        
        # Check sample rows for data quality
        for row_index in range(len(sample_df)):
            row = sample_df.iloc[row_index]
            has_data = False
            row_issues = []
            
            # Check Group A data
            data_a = self.extract_shape_data(row, shape_a, 'A')
            if any(str(v).strip() for v in data_a.values()):
                has_data = True
                # Validate shape-specific data
                shape_issues = self._validate_shape_data(data_a, shape_a, f"Nhóm A")
                row_issues.extend(shape_issues)
            
            # Check Group B data if needed
            if shape_b:
                data_b = self.extract_shape_data(row, shape_b, 'B')
                if any(str(v).strip() for v in data_b.values()):
                    has_data = True
                    shape_issues = self._validate_shape_data(data_b, shape_b, f"Nhóm B")
                    row_issues.extend(shape_issues)
            
            if has_data:
                quality_info['rows_with_data'] += 1
            
            if row_issues:
                quality_info['rows_with_errors'] += 1
                # Only store first 10 error samples to avoid memory issues
                if len(quality_info['data_issues']) < 10:
                    quality_info['data_issues'].append({
                        'row': row_index + 2,  # +2 for Excel row number (1-indexed + header)
                        'issues': row_issues[:3]  # Limit to first 3 issues per row
                    })
        
        return quality_info
    
    def _validate_shape_data(self, data: Dict, shape: str, group_name: str) -> List[str]:
        """Validate data for specific shape - simplified for performance"""
        issues = []
        
        if shape == "Điểm":
            point_input = data.get('point_input', '')
            if point_input:
                coords = point_input.split(',')
                if len(coords) < 2:
                    issues.append(f"{group_name}: Điểm cần ít nhất 2 tọa độ")
        
        elif shape == "Đường thẳng":
            line_A = data.get('line_A1') or data.get('line_A2', '')
            if line_A:
                coords = line_A.split(',')
                if len(coords) != 3:
                    issues.append(f"{group_name}: Đường thẳng cần 3 tọa độ")
        
        elif shape == "Mặt phẳng":
            coeffs = [data.get('plane_a', ''), data.get('plane_b', ''), 
                     data.get('plane_c', ''), data.get('plane_d', '')]
            
            valid_coeffs = sum(1 for coeff in coeffs if coeff and coeff.strip())
            if valid_coeffs == 0:
                issues.append(f"{group_name}: Mặt phẳng cần ít nhất 1 hệ số")
        
        elif shape in ["Đường tròn", "Mặt cầu"]:
            center_key = 'circle_center' if shape == "Đường tròn" else 'sphere_center'
            radius_key = 'circle_radius' if shape == "Đường tròn" else 'sphere_radius'
            
            radius = data.get(radius_key, '')
            if radius:
                try:
                    if float(radius) <= 0:
                        issues.append(f"{group_name}: Bán kính phải > 0")
                except:
                    pass
        
        return issues
    
    def _is_valid_number(self, value: str) -> bool:
        """Check if value is a valid number"""
        try:
            float(value)
            return True
        except:
            return False
    
    def _is_valid_expression(self, value: str) -> bool:
        """Check if value is a valid mathematical expression"""
        if not value or not value.strip():
            return False
        
        # Try simple number first
        if self._is_valid_number(value):
            return True
        
        # Basic expression validation (simplified for performance)
        math_chars = set('0123456789+-*/().\\sqrtsincotan')
        clean_value = value.lower().replace(' ', '')
        return all(c in math_chars for c in clean_value)
