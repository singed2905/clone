from typing import Dict, Any, List, Tuple, Union, Optional
from datetime import datetime
import pandas as pd
import os

from .models import Point2D, Point3D, Line3D, Plane, Circle, Sphere, BaseGeometry
from .mapping_adapter import GeometryMappingAdapter
from .excel_loader import GeometryExcelLoader
from services.excel.excel_processor import ExcelProcessor
from utils.config_loader import config_loader

class GeometryService:
    """Main service for geometry operations - Enhanced with large file support"""
    
    def __init__(self, config: Dict = None):
        self.config = config or {}
        self.mapping_adapter = GeometryMappingAdapter(config)
        self.excel_loader = GeometryExcelLoader(config)
        self.excel_processor = ExcelProcessor(config)  # Enhanced with large file support
        
        # Data storage - matching TL structure
        self.ket_qua_A1 = []  # Line A point coordinates 
        self.ket_qua_X1 = []  # Line A direction vector
        self.ket_qua_N1 = []  # Plane A coefficients
        self.ket_qua_A2 = []  # Line B point coordinates
        self.ket_qua_X2 = []  # Line B direction vector  
        self.ket_qua_N2 = []  # Plane B coefficients
        self.ket_qua_diem_A = []  # Point A coordinates
        self.ket_qua_diem_B = []  # Point B coordinates
        self.ket_qua_duong_tron_A = []  # Circle A parameters
        self.ket_qua_mat_cau_A = []  # Sphere A parameters
        self.ket_qua_duong_tron_B = []  # Circle B parameters 
        self.ket_qua_mat_cau_B = []  # Sphere B parameters
        
        # Store raw data for export
        self.raw_data_A = {}
        self.raw_data_B = {}
        
        # Current state
        self.current_shape_A = ""
        self.current_shape_B = ""
        self.current_operation = ""
        self.kich_thuoc_A = "3"
        self.kich_thuoc_B = "3"
        
        # Geometry data mappings (from TL)
        self.geometry_data = self._init_geometry_data()
        
        # Version config
        self.current_version_config = self._load_version_config()
    
    def _init_geometry_data(self) -> Dict[str, Any]:
        """Initialize geometry data mappings matching TL structure"""
        return {
            "pheptoan_map": {
                "T∆∞∆°ng giao": "qT2",
                "Kho·∫£ng c√°ch": "qT3", 
                "Di·ªán t√≠ch": "qT4",
                "Th·ªÉ t√≠ch": "qT5",
                "PT ƒë∆∞·ªùng th·∫≥ng": "qT6"
            },
            "default_group_a_tcodes": {
                "ƒêi·ªÉm": "T1",
                "ƒê∆∞·ªùng th·∫≥ng": "T4",
                "M·∫∑t ph·∫≥ng": "T7",
                "ƒê∆∞·ªùng tr√≤n": "Tz",
                "M·∫∑t c·∫ßu": "Tj"
            },
            "default_group_b_tcodes": {
                "ƒêi·ªÉm": "T2",
                "ƒê∆∞·ªùng th·∫≥ng": "T5",
                "M·∫∑t ph·∫≥ng": "T8",
                "ƒê∆∞·ªùng tr√≤n": "Tx", 
                "M·∫∑t c·∫ßu": "Tk"
            },
            "operation_tcodes": {
                "Di·ªán t√≠ch": {
                    "group_a": {"ƒê∆∞·ªùng tr√≤n": "T1", "M·∫∑t c·∫ßu": "T4"},
                    "group_b": {"ƒê∆∞·ªùng tr√≤n": "T2", "M·∫∑t c·∫ßu": "T5"}
                },
                "Th·ªÉ t√≠ch": {
                    "group_a": {"M·∫∑t c·∫ßu": "T7"},
                    "group_b": {"M·∫∑t c·∫ßu": "T8"}
                }
            }
        }
    
    def _load_version_config(self) -> Dict[str, Any]:
        """Load version configuration"""
        try:
            if self.config and 'common' in self.config and 'versions' in self.config['common']:
                versions = self.config['common']['versions']
                default_version = versions.get('default_version', 'fx799')
                return config_loader.load_version_config(default_version)
        except Exception as e:
            print(f"Warning: Could not load version config: {e}")
        
        return {"version": "fx799", "prefix": "wj"}
    
    def set_current_shapes(self, shape_A: str, shape_B: str = ""):
        """Set current selected shapes"""
        self.current_shape_A = shape_A
        self.current_shape_B = shape_B
    
    def set_kich_thuoc(self, kich_thuoc_A: str, kich_thuoc_B: str = "3"):
        """Set dimensions"""
        self.kich_thuoc_A = kich_thuoc_A
        self.kich_thuoc_B = kich_thuoc_B
    
    def set_current_operation(self, operation: str):
        """Set current operation"""
        self.current_operation = operation
    
    def cap_nhat_ket_qua(self, chuoi_dau_vao: str, so_tham_so: int = 3, apply_keylog: bool = True) -> List[str]:
        """Update results from input string - matching TL function"""
        if not chuoi_dau_vao:
            return ["" for _ in range(so_tham_so)]

        chuoi_dau_vao = chuoi_dau_vao.replace(" ", "")
        ds = chuoi_dau_vao.split(',')
        while len(ds) < so_tham_so:
            ds.append("0")
        ds = ds[:so_tham_so]

        if apply_keylog:
            ket_qua = [self.mapping_adapter.encode_string(item) for item in ds]
        else:
            ket_qua = ds

        return ket_qua
    
    # ========== GROUP A PROCESSING ==========
    def process_point_A(self, input_data: str) -> List[str]:
        """Process point data for group A"""
        so_chieu = int(self.kich_thuoc_A)
        if so_chieu == 2:
            self.ket_qua_diem_A = self.cap_nhat_ket_qua(input_data, so_tham_so=2)
        else:
            self.ket_qua_diem_A = self.cap_nhat_ket_qua(input_data, so_tham_so=3)
        return self.ket_qua_diem_A
    
    def process_line_A(self, input_A1: str, input_X1: str) -> List[str]:
        """Process line data for group A"""
        self.ket_qua_A1 = self.cap_nhat_ket_qua(input_A1)
        self.ket_qua_X1 = self.cap_nhat_ket_qua(input_X1)
        return self.ket_qua_A1 + self.ket_qua_X1
    
    def process_plane_A(self, coefficients: List[str]) -> List[str]:
        """Process plane data for group A"""
        self.ket_qua_N1 = [self.mapping_adapter.encode_string(coef) for coef in coefficients]
        return self.ket_qua_N1
    
    def process_circle_A(self, center_input: str, radius_input: str) -> List[str]:
        """Process circle data for group A - 2 inputs: center and radius"""
        center_result = self.cap_nhat_ket_qua(center_input, so_tham_so=2)
        radius_result = self.cap_nhat_ket_qua(radius_input, so_tham_so=1)
        self.ket_qua_duong_tron_A = center_result + radius_result
        return self.ket_qua_duong_tron_A
    
    def process_sphere_A(self, center_input: str, radius_input: str) -> List[str]:
        """Process sphere data for group A - 2 inputs: center and radius"""
        center_result = self.cap_nhat_ket_qua(center_input, so_tham_so=3)
        radius_result = self.cap_nhat_ket_qua(radius_input, so_tham_so=1)
        self.ket_qua_mat_cau_A = center_result + radius_result
        return self.ket_qua_mat_cau_A
    
    # ========== GROUP B PROCESSING ==========
    def process_point_B(self, input_data: str) -> List[str]:
        """Process point data for group B"""
        so_chieu = int(self.kich_thuoc_B)
        if so_chieu == 2:
            self.ket_qua_diem_B = self.cap_nhat_ket_qua(input_data, so_tham_so=2)
        else:
            self.ket_qua_diem_B = self.cap_nhat_ket_qua(input_data, so_tham_so=3)
        return self.ket_qua_diem_B
    
    def process_line_B(self, input_A2: str, input_X2: str) -> List[str]:
        """Process line data for group B"""
        self.ket_qua_A2 = self.cap_nhat_ket_qua(input_A2)
        self.ket_qua_X2 = self.cap_nhat_ket_qua(input_X2)
        return self.ket_qua_A2 + self.ket_qua_X2
    
    def process_plane_B(self, coefficients: List[str]) -> List[str]:
        """Process plane data for group B"""
        self.ket_qua_N2 = [self.mapping_adapter.encode_string(coef) for coef in coefficients]
        return self.ket_qua_N2
    
    def process_circle_B(self, center_input: str, radius_input: str) -> List[str]:
        """Process circle data for group B - 2 inputs: center and radius"""
        center_result = self.cap_nhat_ket_qua(center_input, so_tham_so=2)
        radius_result = self.cap_nhat_ket_qua(radius_input, so_tham_so=1)
        self.ket_qua_duong_tron_B = center_result + radius_result
        return self.ket_qua_duong_tron_B
    
    def process_sphere_B(self, center_input: str, radius_input: str) -> List[str]:
        """Process sphere data for group B - 2 inputs: center and radius"""
        center_result = self.cap_nhat_ket_qua(center_input, so_tham_so=3)
        radius_result = self.cap_nhat_ket_qua(radius_input, so_tham_so=1)
        self.ket_qua_mat_cau_B = center_result + radius_result
        return self.ket_qua_mat_cau_B
    
    # ========== MAIN PROCESSING METHODS ==========
    def thuc_thi_A(self, data_dict: Dict[str, str]) -> List[str]:
        """Process group A data based on current shape"""
        shape_type = self.current_shape_A
        self.raw_data_A = data_dict.copy()

        if shape_type == "ƒêi·ªÉm":
            input_data = data_dict.get('point_input', '')
            return self.process_point_A(input_data)

        elif shape_type == "ƒê∆∞·ªùng th·∫≥ng":
            input_A1 = data_dict.get('line_A1', '')
            input_X1 = data_dict.get('line_X1', '')
            return self.process_line_A(input_A1, input_X1)

        elif shape_type == "M·∫∑t ph·∫≥ng":
            coefficients = [
                data_dict.get('plane_a', ''),
                data_dict.get('plane_b', ''),
                data_dict.get('plane_c', ''),
                data_dict.get('plane_d', '')
            ]
            return self.process_plane_A(coefficients)

        elif shape_type == "ƒê∆∞·ªùng tr√≤n":
            center_input = data_dict.get('circle_center', '')
            radius_input = data_dict.get('circle_radius', '')
            return self.process_circle_A(center_input, radius_input)

        elif shape_type == "M·∫∑t c·∫ßu":
            center_input = data_dict.get('sphere_center', '')
            radius_input = data_dict.get('sphere_radius', '')
            return self.process_sphere_A(center_input, radius_input)

        return []
    
    def thuc_thi_B(self, data_dict: Dict[str, str]) -> List[str]:
        """Process group B data based on current shape"""
        if self.current_operation in ["Di·ªán t√≠ch", "Th·ªÉ t√≠ch"]:
            return []

        shape_type = self.current_shape_B
        self.raw_data_B = data_dict.copy()

        if shape_type == "ƒêi·ªÉm":
            input_data = data_dict.get('point_input', '')
            return self.process_point_B(input_data)

        elif shape_type == "ƒê∆∞·ªùng th·∫≥ng":
            input_A2 = data_dict.get('line_A2', '')
            input_X2 = data_dict.get('line_X2', '')
            return self.process_line_B(input_A2, input_X2)

        elif shape_type == "M·∫∑t ph·∫≥ng":
            coefficients = [
                data_dict.get('plane_a', ''),
                data_dict.get('plane_b', ''),
                data_dict.get('plane_c', ''),
                data_dict.get('plane_d', '')
            ]
            return self.process_plane_B(coefficients)

        elif shape_type == "ƒê∆∞·ªùng tr√≤n":
            center_input = data_dict.get('circle_center', '')
            radius_input = data_dict.get('circle_radius', '')
            return self.process_circle_B(center_input, radius_input)

        elif shape_type == "M·∫∑t c·∫ßu":
            center_input = data_dict.get('sphere_center', '')
            radius_input = data_dict.get('sphere_radius', '')
            return self.process_sphere_B(center_input, radius_input)

        return []
    
    def thuc_thi_tat_ca(self, data_dict_A: Dict[str, str], data_dict_B: Dict[str, str] = None) -> Tuple[List[str], List[str]]:
        """Process all groups"""
        if data_dict_B is None:
            data_dict_B = {}
        
        result_A = self.thuc_thi_A(data_dict_A)
        result_B = self.thuc_thi_B(data_dict_B)
        return result_A, result_B
    
    # ========== EXCEL INTEGRATION - ENHANCED FOR LARGE FILES ==========
    def process_excel_batch(self, file_path: str, shape_a: str, shape_b: str, 
                           operation: str, dimension_a: str, dimension_b: str,
                           output_path: str = None, progress_callback: callable = None) -> Tuple[List[str], str, int, int]:
        """Process Excel file - Auto-detect large files and use appropriate method"""
        try:
            # Detect if this is a large file
            is_large, file_info = self.excel_processor.is_large_file(file_path)
            
            if is_large:
                print(f"üî• LARGE FILE DETECTED: {file_info['file_size_mb']:.1f}MB, {file_info['estimated_rows']:,} rows")
                print(f"üöÄ Switching to Large File Processor...")
                
                # Generate output path if not provided
                if not output_path:
                    original_name = os.path.splitext(os.path.basename(file_path))[0]
                    output_path = f"{original_name}_large_encoded_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    output_path = os.path.join(os.path.dirname(file_path), output_path)
                
                # Use large file processor
                success_count, error_count, output_file = self.excel_processor.process_large_excel_file(
                    file_path, shape_a, shape_b, operation, dimension_a, dimension_b,
                    output_path, progress_callback
                )
                
                # Return format matching original method
                return [], output_file, success_count, error_count
            
            else:
                # Use normal processing for smaller files
                return self._process_excel_normal(file_path, shape_a, shape_b, operation, 
                                                dimension_a, dimension_b, output_path, progress_callback)
            
        except Exception as e:
            raise Exception(f"L·ªói x·ª≠ l√Ω Excel: {str(e)}")
    
    def _process_excel_normal(self, file_path: str, shape_a: str, shape_b: str,
                            operation: str, dimension_a: str, dimension_b: str,
                            output_path: str = None, progress_callback: callable = None) -> Tuple[List[str], str, int, int]:
        """Normal Excel processing for smaller files"""
        try:
            # Read and validate Excel file
            df = self.excel_processor.read_excel_data(file_path)
            is_valid, missing_cols = self.excel_processor.validate_excel_structure(df, shape_a, shape_b)

            if not is_valid:
                raise Exception(f"Thi·∫øu c√°c c·ªôt: {', '.join(missing_cols)}")

            encoded_results = []
            processed_count = 0
            error_count = 0
            total_rows = len(df)

            # Process each row
            for index, row in df.iterrows():
                try:
                    # Set current state for this row
                    self.set_current_shapes(shape_a, shape_b)
                    self.set_kich_thuoc(dimension_a, dimension_b)
                    self.current_operation = operation

                    # Extract data for both groups
                    data_a = self.excel_processor.extract_shape_data(row, shape_a, 'A')
                    data_b = self.excel_processor.extract_shape_data(row, shape_b, 'B') if shape_b else {}

                    # Process data
                    self.thuc_thi_tat_ca(data_a, data_b)
                    result = self.generate_final_result()

                    encoded_results.append(result)
                    processed_count += 1
                    
                    # Update progress if callback provided
                    if progress_callback and (processed_count % 10 == 0 or processed_count == total_rows):
                        progress = (processed_count / total_rows) * 100
                        progress_callback(progress, processed_count, total_rows, error_count)

                except Exception as e:
                    # Log error but continue with next row
                    encoded_results.append(f"L·ªñI: {str(e)}")
                    error_count += 1
                    print(f"L·ªói d√≤ng {index + 1}: {str(e)}")

            # Generate output path if not provided
            if not output_path:
                original_name = os.path.splitext(os.path.basename(file_path))[0]
                output_path = f"{original_name}_encoded_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                output_path = os.path.join(os.path.dirname(file_path), output_path)

            # Export results
            output_file = self.excel_processor.export_results(df, encoded_results, output_path)

            return encoded_results, output_file, processed_count, error_count

        except Exception as e:
            raise Exception(f"L·ªói x·ª≠ l√Ω file Excel th√¥ng th∆∞·ªùng: {str(e)}")
    
    def process_excel_batch_chunked(self, file_path: str, shape_a: str, shape_b: str,
                                  operation: str, dimension_a: str, dimension_b: str,
                                  chunksize: int = 1000, progress_callback: callable = None) -> Tuple[List[str], str, int, int]:
        """Process Excel file in chunks - Enhanced for large files"""
        try:
            # Check if this is a very large file
            is_large, file_info = self.excel_processor.is_large_file(file_path)
            
            if is_large:
                print(f"üî• VERY LARGE FILE - Using specialized chunked processor")
                
                # Generate output path
                original_name = os.path.splitext(os.path.basename(file_path))[0]
                output_path = f"{original_name}_chunked_large_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                output_path = os.path.join(os.path.dirname(file_path), output_path)
                
                # Use large file processor with chunking
                success_count, error_count, output_file = self.excel_processor.process_large_excel_file(
                    file_path, shape_a, shape_b, operation, dimension_a, dimension_b,
                    output_path, progress_callback
                )
                
                return [], output_file, success_count, error_count
            
            else:
                # Normal chunked processing for medium files
                return self._process_excel_chunked_normal(file_path, shape_a, shape_b, operation,
                                                        dimension_a, dimension_b, chunksize, progress_callback)
                
        except Exception as e:
            raise Exception(f"L·ªói x·ª≠ l√Ω Excel chunked: {str(e)}")
    
    def _process_excel_chunked_normal(self, file_path: str, shape_a: str, shape_b: str,
                                    operation: str, dimension_a: str, dimension_b: str,
                                    chunksize: int, progress_callback: callable = None) -> Tuple[List[str], str, int, int]:
        """Normal chunked processing for medium files"""
        try:
            # Get total rows for progress calculation
            total_rows = self.excel_processor.get_total_rows(file_path)
            processed_count = 0
            error_count = 0
            all_results = []
            
            # Validate structure first
            df_sample = self.excel_processor.read_excel_data(file_path)
            is_valid, missing_cols = self.excel_processor.validate_excel_structure(df_sample, shape_a, shape_b)
            if not is_valid:
                raise Exception(f"Thi·∫øu c√°c c·ªôt: {', '.join(missing_cols)}")

            # Process in chunks
            chunk_iterator = self.excel_processor.read_excel_data_chunked(file_path, chunksize)

            for chunk_idx, chunk_df in enumerate(chunk_iterator):
                chunk_results = []

                # Process each row in chunk
                for index, row in chunk_df.iterrows():
                    try:
                        # Set current state
                        self.set_current_shapes(shape_a, shape_b)
                        self.set_kich_thuoc(dimension_a, dimension_b)
                        self.current_operation = operation

                        # Extract and process data
                        data_a = self.excel_processor.extract_shape_data(row, shape_a, 'A')
                        data_b = self.excel_processor.extract_shape_data(row, shape_b, 'B') if shape_b else {}

                        self.thuc_thi_tat_ca(data_a, data_b)
                        result = self.generate_final_result()

                        chunk_results.append(result)
                        processed_count += 1

                    except Exception as e:
                        chunk_results.append(f"L·ªñI: {str(e)}")
                        error_count += 1

                    # Update progress every 10 rows
                    if progress_callback and processed_count % 10 == 0:
                        progress = (processed_count / total_rows) * 100 if total_rows > 0 else 0
                        progress_callback(progress, processed_count, total_rows, error_count)

                all_results.extend(chunk_results)

                # Memory cleanup after each chunk
                del chunk_df
                import gc
                gc.collect()

            # Export final results
            original_name = os.path.splitext(os.path.basename(file_path))[0]
            output_path = f"{original_name}_chunked_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(os.path.dirname(file_path), output_path)

            final_df = self.excel_processor.read_excel_data(file_path)
            output_file = self.excel_processor.export_results(final_df, all_results, output_path)

            return all_results, output_file, processed_count, error_count

        except Exception as e:
            raise Exception(f"L·ªói x·ª≠ l√Ω chunked th√¥ng th∆∞·ªùng: {str(e)}")
    
    def validate_excel_file(self, file_path: str, shape_a: str, shape_b: str = None) -> Dict[str, Any]:
        """Comprehensive Excel file validation - Enhanced for large files"""
        try:
            # Basic file validation
            if not os.path.exists(file_path):
                return {'valid': False, 'error': 'File kh√¥ng t·ªìn t·∫°i'}
            
            # Check if large file
            is_large, large_file_info = self.excel_processor.is_large_file(file_path)
            
            if is_large:
                print(f"üìÑ Large file validation: {large_file_info['file_size_mb']:.1f}MB")
                
                # Use large file validation
                validation_result = self.excel_processor.validate_large_file_structure(file_path, shape_a, shape_b)
                
                return {
                    'valid': validation_result['valid'],
                    'is_large_file': True,
                    'file_info': {
                        'file_name': os.path.basename(file_path),
                        'file_size_mb': large_file_info['file_size_mb'],
                        'estimated_rows': large_file_info['estimated_rows'],
                        'recommended_chunk_size': large_file_info['recommended_chunk_size']
                    },
                    'structure_issues': validation_result.get('missing_columns', []),
                    'quality_issues': {'note': 'Large file - quality check limited to structure only'},
                    'ready_for_processing': validation_result['valid'],
                    'processing_recommendation': 'Use large file processor with chunking'
                }
            
            else:
                # Normal file validation
                file_info = self.excel_processor.get_file_info(file_path)
                df = self.excel_processor.read_excel_data(file_path)
                structure_valid, missing_cols = self.excel_processor.validate_excel_structure(df, shape_a, shape_b)
                quality_info = self.excel_processor.validate_data_quality(df, shape_a, shape_b)
                
                return {
                    'valid': structure_valid and quality_info['valid'],
                    'is_large_file': False,
                    'file_info': file_info,
                    'structure_issues': missing_cols,
                    'quality_issues': quality_info,
                    'ready_for_processing': structure_valid and quality_info['rows_with_data'] > 0
                }
            
        except Exception as e:
            return {'valid': False, 'error': f'L·ªói ki·ªÉm tra file: {str(e)}'}
    
    def export_single_result(self, output_path: str = None) -> str:
        """Export current single result to Excel"""
        try:
            if output_path is None:
                output_path = f"geometry_single_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                output_path = os.path.join(os.getcwd(), output_path)

            # Prepare comprehensive export data
            data = self._prepare_comprehensive_export_data()
            df = pd.DataFrame(data)

            # Create Excel with multiple sheets
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Geometry Data', index=False)

                # Add summary sheet
                summary_data = self._prepare_summary_data()
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format worksheets
                self._format_export_worksheets(writer, df, summary_df)

            return output_path

        except ImportError:
            raise Exception("Th∆∞ vi·ªán openpyxl ch∆∞a ƒë∆∞·ª£c c√†i ƒë·∫∑t. Vui l√≤ng c√†i ƒë·∫∑t b·∫±ng l·ªánh: pip install openpyxl")
        except Exception as e:
            raise Exception(f"L·ªói xu·∫•t Excel: {str(e)}")
    
    def _prepare_comprehensive_export_data(self):
        """Prepare comprehensive data for Excel export - matching TL structure"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Base data
        data = {
            "Th·ªùi gian": [timestamp],
            "Ph√©p to√°n": [self.current_operation],
            "ƒê·ªëi t∆∞·ª£ng A": [self.current_shape_A],
            "ƒê·ªëi t∆∞·ª£ng B": [
                self.current_shape_B if self.current_operation not in ["Di·ªán t√≠ch", "Th·ªÉ t√≠ch"] else "Kh√¥ng c√≥"],
            "K·∫øt qu·∫£ m√£ h√≥a": [self.generate_final_result()],
            "K√≠ch th∆∞·ªõc A": [self.kich_thuoc_A],
            "K√≠ch th∆∞·ªõc B": [self.kich_thuoc_B]
        }

        # Add detailed data based on shape types
        self._add_detailed_export_data(data, "A", self.current_shape_A, self.raw_data_A)

        if self.current_operation not in ["Di·ªán t√≠ch", "Th·ªÉ t√≠ch"]:
            self._add_detailed_export_data(data, "B", self.current_shape_B, self.raw_data_B)

        return data
    
    def _add_detailed_export_data(self, data, group, shape_type, raw_data):
        """Add detailed export data for specific group and shape type - matching TL"""
        prefix = f"Nh√≥m {group} - "

        if shape_type == "ƒêi·ªÉm":
            point_input = raw_data.get('point_input', '')
            coords = point_input.split(',') if point_input else []
            
            data[f"{prefix}To·∫° ƒë·ªô X"] = [coords[0] if len(coords) > 0 else ""]
            data[f"{prefix}To·∫° ƒë·ªô Y"] = [coords[1] if len(coords) > 1 else ""]
            if group == "A":
                data[f"{prefix}To·∫° ƒë·ªô X (m√£ h√≥a)"] = [self.ket_qua_diem_A[0] if len(self.ket_qua_diem_A) > 0 else ""]
                data[f"{prefix}To·∫° ƒë·ªô Y (m√£ h√≥a)"] = [self.ket_qua_diem_A[1] if len(self.ket_qua_diem_A) > 1 else ""]
                if int(self.kich_thuoc_A) == 3:
                    data[f"{prefix}To·∫° ƒë·ªô Z"] = [coords[2] if len(coords) > 2 else ""]
                    data[f"{prefix}To·∫° ƒë·ªô Z (m√£ h√≥a)"] = [self.ket_qua_diem_A[2] if len(self.ket_qua_diem_A) > 2 else ""]
            else:
                data[f"{prefix}To·∫° ƒë·ªô X (m√£ h√≥a)"] = [self.ket_qua_diem_B[0] if len(self.ket_qua_diem_B) > 0 else ""]
                data[f"{prefix}To·∫° ƒë·ªô Y (m√£ h√≥a)"] = [self.ket_qua_diem_B[1] if len(self.ket_qua_diem_B) > 1 else ""]
                if int(self.kich_thuoc_B) == 3:
                    data[f"{prefix}To·∫° ƒë·ªô Z"] = [coords[2] if len(coords) > 2 else ""]
                    data[f"{prefix}To·∫° ƒë·ªô Z (m√£ h√≥a)"] = [self.ket_qua_diem_B[2] if len(self.ket_qua_diem_B) > 2 else ""]

        elif shape_type == "ƒê∆∞·ªùng th·∫≥ng":
            line_A = raw_data.get('line_A1') or raw_data.get('line_A2', '')
            line_X = raw_data.get('line_X1') or raw_data.get('line_X2', '')
            
            point_coords = line_A.split(',') if line_A else []
            vector_coords = line_X.split(',') if line_X else []
            
            data[f"{prefix}ƒêi·ªÉm A"] = [point_coords[0] if len(point_coords) > 0 else ""]
            data[f"{prefix}ƒêi·ªÉm B"] = [point_coords[1] if len(point_coords) > 1 else ""]
            data[f"{prefix}ƒêi·ªÉm C"] = [point_coords[2] if len(point_coords) > 2 else ""]
            data[f"{prefix}Vector X"] = [vector_coords[0] if len(vector_coords) > 0 else ""]
            data[f"{prefix}Vector Y"] = [vector_coords[1] if len(vector_coords) > 1 else ""]
            data[f"{prefix}Vector Z"] = [vector_coords[2] if len(vector_coords) > 2 else ""]

        # Add other shape types as needed...
    
    def _prepare_summary_data(self):
        """Prepare summary data for Excel export"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        summary = {
            "Th·ªùi gian xu·∫•t": [timestamp],
            "T·ªïng s·ªë ƒë·ªëi t∆∞·ª£ng": ["2" if self.current_operation not in ["Di·ªán t√≠ch", "Th·ªÉ t√≠ch"] else "1"],
            "Ph√©p to√°n th·ª±c hi·ªán": [self.current_operation],
            "ƒê·ªëi t∆∞·ª£ng ch√≠nh": [self.current_shape_A],
            "ƒê·ªëi t∆∞·ª£ng ph·ª•": [
                self.current_shape_B if self.current_operation not in ["Di·ªán t√≠ch", "Th·ªÉ t√≠ch"] else "Kh√¥ng c√≥"],
            "Tr·∫°ng th√°i": ["ƒê√£ x·ª≠ l√Ω v√† m√£ h√≥a"],
            "ƒê·ªô d√†i k·∫øt qu·∫£": [len(self.generate_final_result())],
            "Ghi ch√∫": ["D·ªØ li·ªáu ƒë√£ ƒë∆∞·ª£c m√£ h√≥a theo quy t·∫Øc mapping"]
        }

        return summary
    
    def _format_export_worksheets(self, writer, main_df, summary_df):
        """Format export worksheets"""
        try:
            # Format main sheet
            main_ws = writer.sheets['Geometry Data']
            self.excel_processor._format_results_worksheet(main_ws, main_df)
            
            # Format summary sheet
            summary_ws = writer.sheets['Summary']
            from openpyxl.styles import Font, PatternFill
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
            
            for col in range(1, len(summary_df.columns) + 1):
                cell = summary_ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                
        except Exception as e:
            print(f"Warning: Could not format export worksheets: {e}")
    
    def create_template_for_shapes(self, shape_a: str, shape_b: str = None, output_path: str = None) -> str:
        """Create Excel template for specific shape combination"""
        if output_path is None:
            shapes_name = f"{shape_a}" + (f"_{shape_b}" if shape_b else "")
            output_path = f"template_{shapes_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(os.getcwd(), output_path)
        
        return self.excel_processor.create_geometry_template(shape_a, shape_b, output_path)
    
    def generate_final_result(self) -> str:
        """Generate the final encoded string - matching TL behavior"""
        if not self.current_shape_A or not self.current_operation:
            return ""

        pheptoan_code = self.geometry_data["pheptoan_map"].get(self.current_operation, self.current_operation)

        # Get T-code mappings
        tcodeA = self._get_tcode_mapping("A", self.current_shape_A)

        # For Area and Volume, don't include group B
        if self.current_operation in ["Di·ªán t√≠ch", "Th·ªÉ t√≠ch"]:
            tenB_code = ""
            gia_tri_B = ""
            tcodeB = ""
        else:
            tcodeB = self._get_tcode_mapping("B", self.current_shape_B)
            tenB_code = self._get_shape_code_B(self.current_shape_B)
            gia_tri_B = self._get_encoded_values_B()

        tenA_code = self._get_shape_code_A(self.current_shape_A)
        gia_tri_A = self._get_encoded_values_A()
        prefix = self.current_version_config.get("prefix", "wj")
        
        # Build final result - matching TL format
        if self.current_operation in ["Di·ªán t√≠ch", "Th·ªÉ t√≠ch"]:
            ket_qua = f"{prefix}{tenA_code}{gia_tri_A}C{pheptoan_code}{tcodeA}="
        else:
            ket_qua = f"{prefix}{tenA_code}{gia_tri_A}C{tenB_code}{gia_tri_B}C{pheptoan_code}{tcodeA}R{tcodeB}="

        return ket_qua
    
    def _get_tcode_mapping(self, group: str, shape: str) -> str:
        """Get T-code mapping for shape - matching TL logic"""
        pheptoan = self.current_operation

        if pheptoan in self.geometry_data["operation_tcodes"]:
            operation_map = self.geometry_data["operation_tcodes"][pheptoan]
            if group == "A" and shape in operation_map["group_a"]:
                return operation_map["group_a"][shape]
            elif group == "B" and shape in operation_map["group_b"]:
                return operation_map["group_b"][shape]

        # Return default mapping if no operation-specific mapping found
        if group == "A":
            return self.geometry_data["default_group_a_tcodes"].get(shape, "T0")
        else:
            return self.geometry_data["default_group_b_tcodes"].get(shape, "T0")
    
    def _get_shape_code_A(self, shape: str) -> str:
        """Get shape code for group A - matching TL logic"""
        if shape == "ƒêi·ªÉm" and self.kich_thuoc_A == "2":
            return "112"
        elif shape == "ƒêi·ªÉm" and self.kich_thuoc_A == "3":
            return "113"
        elif shape == "ƒê∆∞·ªùng th·∫≥ng":
            return "21"
        elif shape == "M·∫∑t ph·∫≥ng":
            return "31"
        elif shape == "ƒê∆∞·ªùng tr√≤n":
            return "41"
        elif shape == "M·∫∑t c·∫ßu":
            return "51"
        else:
            return "00"
    
    def _get_shape_code_B(self, shape: str) -> str:
        """Get shape code for group B - matching TL logic"""
        if shape == "ƒêi·ªÉm" and self.kich_thuoc_B == "2":
            return "qT11T122"
        elif shape == "ƒêi·ªÉm" and self.kich_thuoc_B == "3":
            return "qT11T123"
        elif shape == "ƒê∆∞·ªùng th·∫≥ng":
            return "qT12T12"
        elif shape == "M·∫∑t ph·∫≥ng":
            return "qT13T12"
        elif shape == "ƒê∆∞·ªùng tr√≤n":
            return "qT14T12"
        elif shape == "M·∫∑t c·∫ßu":
            return "qT15T12"
        else:
            return "qT00T12"
    
    def _get_encoded_values_A(self) -> str:
        """Get encoded values for group A - matching TL format"""
        shape = self.current_shape_A

        if shape == "ƒêi·ªÉm":
            so_chieu = int(self.kich_thuoc_A)
            if so_chieu == 2:
                x_encoded = self.ket_qua_diem_A[0] if len(self.ket_qua_diem_A) > 0 else ""
                y_encoded = self.ket_qua_diem_A[1] if len(self.ket_qua_diem_A) > 1 else ""
                return f"{x_encoded}={y_encoded}="
            else:
                x_encoded = self.ket_qua_diem_A[0] if len(self.ket_qua_diem_A) > 0 else ""
                y_encoded = self.ket_qua_diem_A[1] if len(self.ket_qua_diem_A) > 1 else ""
                z_encoded = self.ket_qua_diem_A[2] if len(self.ket_qua_diem_A) > 2 else ""
                return f"{x_encoded}={y_encoded}={z_encoded}="

        elif shape == "ƒê∆∞·ªùng th·∫≥ng":
            A_encoded = self.ket_qua_A1[0] if len(self.ket_qua_A1) > 0 else ""
            X_encoded = self.ket_qua_X1[0] if len(self.ket_qua_X1) > 0 else ""
            B_encoded = self.ket_qua_A1[1] if len(self.ket_qua_A1) > 1 else ""
            Y_encoded = self.ket_qua_X1[1] if len(self.ket_qua_X1) > 1 else ""
            C_encoded = self.ket_qua_A1[2] if len(self.ket_qua_A1) > 2 else ""
            Z_encoded = self.ket_qua_X1[2] if len(self.ket_qua_X1) > 2 else ""
            return f"{A_encoded}={X_encoded}={B_encoded}={Y_encoded}={C_encoded}={Z_encoded}="

        elif shape == "M·∫∑t ph·∫≥ng":
            N1_encoded = self.ket_qua_N1[0] if len(self.ket_qua_N1) > 0 else ""
            N2_encoded = self.ket_qua_N1[1] if len(self.ket_qua_N1) > 1 else ""
            N3_encoded = self.ket_qua_N1[2] if len(self.ket_qua_N1) > 2 else ""
            N4_encoded = self.ket_qua_N1[3] if len(self.ket_qua_N1) > 3 else ""
            return f"{N1_encoded}={N2_encoded}={N3_encoded}={N4_encoded}="

        elif shape == "ƒê∆∞·ªùng tr√≤n":
            A1_encoded = self.ket_qua_duong_tron_A[0] if len(self.ket_qua_duong_tron_A) > 0 else ""
            A2_encoded = self.ket_qua_duong_tron_A[1] if len(self.ket_qua_duong_tron_A) > 1 else ""
            A3_encoded = self.ket_qua_duong_tron_A[2] if len(self.ket_qua_duong_tron_A) > 2 else ""
            return f"{A1_encoded}={A2_encoded}={A3_encoded}="

        elif shape == "M·∫∑t c·∫ßu":
            A1_encoded = self.ket_qua_mat_cau_A[0] if len(self.ket_qua_mat_cau_A) > 0 else ""
            A2_encoded = self.ket_qua_mat_cau_A[1] if len(self.ket_qua_mat_cau_A) > 1 else ""
            A3_encoded = self.ket_qua_mat_cau_A[2] if len(self.ket_qua_mat_cau_A) > 2 else ""
            A4_encoded = self.ket_qua_mat_cau_A[3] if len(self.ket_qua_mat_cau_A) > 3 else ""
            return f"{A1_encoded}={A2_encoded}={A3_encoded}={A4_encoded}="

        return ""
    
    def _get_encoded_values_B(self) -> str:
        """Get encoded values for group B - matching TL format"""
        if self.current_operation in ["Di·ªán t√≠ch", "Th·ªÉ t√≠ch"]:
            return ""

        shape = self.current_shape_B

        if shape == "ƒêi·ªÉm":
            so_chieu = int(self.kich_thuoc_B)
            if so_chieu == 2:
                x_encoded = self.ket_qua_diem_B[0] if len(self.ket_qua_diem_B) > 0 else ""
                y_encoded = self.ket_qua_diem_B[1] if len(self.ket_qua_diem_B) > 1 else ""
                return f"{x_encoded}={y_encoded}="
            else:
                x_encoded = self.ket_qua_diem_B[0] if len(self.ket_qua_diem_B) > 0 else ""
                y_encoded = self.ket_qua_diem_B[1] if len(self.ket_qua_diem_B) > 1 else ""
                z_encoded = self.ket_qua_diem_B[2] if len(self.ket_qua_diem_B) > 2 else ""
                return f"{x_encoded}={y_encoded}={z_encoded}="

        elif shape == "ƒê∆∞·ªùng th·∫≥ng":
            A_encoded = self.ket_qua_A2[0] if len(self.ket_qua_A2) > 0 else ""
            X_encoded = self.ket_qua_X2[0] if len(self.ket_qua_X2) > 0 else ""
            B_encoded = self.ket_qua_A2[1] if len(self.ket_qua_A2) > 1 else ""
            Y_encoded = self.ket_qua_X2[1] if len(self.ket_qua_X2) > 1 else ""
            C_encoded = self.ket_qua_A2[2] if len(self.ket_qua_A2) > 2 else ""
            Z_encoded = self.ket_qua_X2[2] if len(self.ket_qua_X2) > 2 else ""
            return f"{A_encoded}={X_encoded}={B_encoded}={Y_encoded}={C_encoded}={Z_encoded}="

        elif shape == "M·∫∑t ph·∫≥ng":
            N5_encoded = self.ket_qua_N2[0] if len(self.ket_qua_N2) > 0 else ""
            N6_encoded = self.ket_qua_N2[1] if len(self.ket_qua_N2) > 1 else ""
            N7_encoded = self.ket_qua_N2[2] if len(self.ket_qua_N2) > 2 else ""
            N8_encoded = self.ket_qua_N2[3] if len(self.ket_qua_N2) > 3 else ""
            return f"{N5_encoded}={N6_encoded}={N7_encoded}={N8_encoded}="

        elif shape == "ƒê∆∞·ªùng tr√≤n":
            B1_encoded = self.ket_qua_duong_tron_B[0] if len(self.ket_qua_duong_tron_B) > 0 else ""
            B2_encoded = self.ket_qua_duong_tron_B[1] if len(self.ket_qua_duong_tron_B) > 1 else ""
            B3_encoded = self.ket_qua_duong_tron_B[2] if len(self.ket_qua_duong_tron_B) > 2 else ""
            return f"{B1_encoded}={B2_encoded}={B3_encoded}="

        elif shape == "M·∫∑t c·∫ßu":
            B1_encoded = self.ket_qua_mat_cau_B[0] if len(self.ket_qua_mat_cau_B) > 0 else ""
            B2_encoded = self.ket_qua_mat_cau_B[1] if len(self.ket_qua_mat_cau_B) > 1 else ""
            B3_encoded = self.ket_qua_mat_cau_B[2] if len(self.ket_qua_mat_cau_B) > 2 else ""
            B4_encoded = self.ket_qua_mat_cau_B[3] if len(self.ket_qua_mat_cau_B) > 3 else ""
            return f"{B1_encoded}={B2_encoded}={B3_encoded}={B4_encoded}="

        return ""
    
    # ========== CONVENIENCE METHODS ==========
    def get_available_shapes(self) -> List[str]:
        """Get list of available geometric shapes"""
        return ["ƒêi·ªÉm", "ƒê∆∞·ªùng th·∫≥ng", "M·∫∑t ph·∫≥ng", "ƒê∆∞·ªùng tr√≤n", "M·∫∑t c·∫ßu"]
    
    def get_available_operations(self) -> List[str]:
        """Get list of available operations"""
        return ["T∆∞∆°ng giao", "Kho·∫£ng c√°ch", "Di·ªán t√≠ch", "Th·ªÉ t√≠ch", "PT ƒë∆∞·ªùng th·∫≥ng"]
    
    def update_dropdown_options(self, operation: str) -> List[str]:
        """Update dropdown options based on selected operation"""
        self.current_operation = operation
        if operation == "Kho·∫£ng c√°ch":
            return ["ƒêi·ªÉm", "ƒê∆∞·ªùng th·∫≥ng", "M·∫∑t ph·∫≥ng"]
        elif operation == "Di·ªán t√≠ch":
            return ["ƒê∆∞·ªùng tr√≤n", "M·∫∑t c·∫ßu"]
        elif operation == "Th·ªÉ t√≠ch":
            return ["M·∫∑t c·∫ßu"]
        elif operation == "PT ƒë∆∞·ªùng th·∫≥ng":
            return ["ƒêi·ªÉm"]
        else:
            return self.get_available_shapes()
    
    def get_result_summary(self) -> Dict[str, Any]:
        """Get summary of current processing results"""
        return {
            "operation": self.current_operation,
            "shape_A": self.current_shape_A,
            "shape_B": self.current_shape_B,
            "dimensions_A": self.kich_thuoc_A,
            "dimensions_B": self.kich_thuoc_B,
            "encoded_result": self.generate_final_result(),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    
    # ========== ENHANCED EXCEL METHODS ==========
    def get_excel_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get comprehensive Excel file information - Enhanced with large file detection"""
        return self.excel_processor.get_file_info(file_path)
    
    def validate_excel_file_for_geometry(self, file_path: str, shape_a: str, shape_b: str = None) -> Dict[str, Any]:
        """Validate Excel file for geometry processing - Enhanced for large files"""
        return self.validate_excel_file(file_path, shape_a, shape_b)
    
    def create_excel_template_for_geometry(self, shape_a: str, shape_b: str = None, output_path: str = None) -> str:
        """Create Excel template for specific geometry shapes"""
        return self.create_template_for_shapes(shape_a, shape_b, output_path)
    
    def get_large_file_processing_info(self, file_path: str) -> Dict[str, Any]:
        """Get information about large file processing capabilities"""
        try:
            is_large, file_info = self.excel_processor.is_large_file(file_path)
            
            return {
                'is_large_file': is_large,
                'file_info': file_info,
                'processing_mode': 'large_file_processor' if is_large else 'normal_processor',
                'memory_optimization': is_large,
                'streaming_mode': is_large,
                'recommended_chunk_size': file_info.get('recommended_chunk_size', 1000)
            }
        except Exception as e:
            return {'error': f'Kh√¥ng th·ªÉ ph√¢n t√≠ch file: {str(e)}'}
